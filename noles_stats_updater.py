"""
Beyond Howser — Stats Updater
==================================
Pulls current season stats from the MLB Stats API for all players
on the Beyond Howser roster, then:
  1. Updates a "2026 Stats" sheet in noles_in_the_pros.xlsx
  2. Generates/refreshes noles_dashboard.html

Run manually or via scheduled task. Requires: openpyxl, requests
  pip install openpyxl requests
"""

import json
import os
import requests
from datetime import datetime
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR        = Path(__file__).parent
EXCEL_PATH      = BASE_DIR / "noles_in_the_pros.xlsx"
CACHE_PATH      = BASE_DIR / "player_id_cache.json"
NEWS_CACHE_PATH = BASE_DIR / "news_cache.json"
HTML_PATH       = BASE_DIR / "noles_dashboard.html"
SEASON          = datetime.now().year
MLB_API    = "https://statsapi.mlb.com/api/v1"

# sport IDs: 1=MLB, 11=AAA, 12=AA, 13=High-A, 14=Low-A, 15=Rookie/Short, 16=Complex
SPORT_IDS  = "1,11,12,13,14,15,16"
LEVEL_SPORT_ID = {
    "MLB":         1,
    "AAA":         11,
    "AA":          12,
    "High-A":      13,
    "Low-A":       14,
    "Rookie":      15,
    "Independent": None,
    "60-Day IL":   None,
    "7-Day IL":    None,
    "Released":    None,
}

# Levels that skip API stat lookups entirely
NO_STATS_LEVELS = {"Released", "60-Day IL", "7-Day IL", "Independent"}

HEADERS = {"User-Agent": "BeyondHowser/1.0"}

# ── Player ID Cache ───────────────────────────────────────────────────────────
def load_cache():
    if CACHE_PATH.exists():
        with open(CACHE_PATH) as f:
            return json.load(f)
    return {}

def save_cache(cache):
    with open(CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)

def find_player_id(name: str, cache: dict) -> int | None:
    """Search MLB Stats API for a player by name. Returns personId or None."""
    if name in cache:
        return cache[name]
    try:
        r = requests.get(
            f"{MLB_API}/people/search",
            params={"names": name, "sportIds": SPORT_IDS},
            headers=HEADERS, timeout=10
        )
        r.raise_for_status()
        results = r.json().get("people", [])
        if results:
            pid = results[0]["id"]
            cache[name] = pid
            save_cache(cache)
            print(f"  ✓ Found {name} → ID {pid}")
            return pid
        else:
            print(f"  ✗ Not found in MLB API: {name} (likely independent league)")
            cache[name] = None
            save_cache(cache)
            return None
    except Exception as e:
        print(f"  ! Error looking up {name}: {e}")
        return None

# ── Stats Fetching ────────────────────────────────────────────────────────────
def get_player_stats(person_id: int, season: int, level: str = "") -> dict:
    """Fetch hitting and pitching stats using the player's level to target
    the correct sport ID (e.g. AAA=11, AA=12) for a clean single API call."""
    stats = {"hitting": {}, "pitching": {}, "season_used": season}
    sport_id = LEVEL_SPORT_ID.get(level)
    if sport_id is None:
        return stats  # Independent league or unknown — no API data available
    for group in ("hitting", "pitching"):
        try:
            r = requests.get(
                f"{MLB_API}/people/{person_id}/stats",
                params={"stats": "season", "season": season,
                        "group": group, "sportId": sport_id},
                headers=HEADERS, timeout=10
            )
            r.raise_for_status()
            splits = r.json().get("stats", [])
            if splits and splits[0].get("splits"):
                stats[group] = splits[0]["splits"][0].get("stat", {})
        except Exception as e:
            print(f"    ! Stats error for {person_id} ({group}): {e}")
    return stats

# ── Read Roster from Excel ────────────────────────────────────────────────────
def read_roster() -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Roster"]
    players = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name, pos, org, level, team = row[0], row[1], row[2], row[3], row[4]
        draft_year  = row[5] if len(row) > 5 else None
        draft_round = row[6] if len(row) > 6 else None
        draft_pick  = row[7] if len(row) > 7 else None
        milb_url    = row[9] if len(row) > 9 else None
        notes       = row[10] if len(row) > 10 else None
        base_level  = row[11] if len(row) > 11 else None
        if name:
            draft_str = ""
            try:
                if draft_year and draft_round and str(draft_year).strip('—'):
                    pick_str = f" (Pick #{int(float(draft_pick))})" if draft_pick and str(draft_pick).strip('—').strip() else ""
                    draft_str = f"{int(float(draft_year))} · Rd {int(float(draft_round))}{pick_str}"
            except (ValueError, TypeError):
                draft_str = ""
            players.append({
                "name": name, "position": pos or "",
                "org": org or "", "level": level or "",
                "team": team or "", "milb_url": milb_url or "",
                "draft": draft_str, "notes": str(notes).strip() if notes else "",
                "base_level": str(base_level).strip() if base_level else ""
            })
    return players

# ── Format Stats for Display ─────────────────────────────────────────────────
def format_hitting(s: dict) -> dict:
    def fmt(val, decimals=0):
        if val is None: return "—"
        try:
            v = float(val)
            if decimals: return f"{v:.{decimals}f}".lstrip("0") or ".000"
            return str(int(v))
        except: return str(val)

    return {
        "G":   fmt(s.get("gamesPlayed")),
        "AB":  fmt(s.get("atBats")),
        "H":   fmt(s.get("hits")),
        "AVG": fmt(s.get("avg"), 3),
        "HR":  fmt(s.get("homeRuns")),
        "RBI": fmt(s.get("rbi")),
        "R":   fmt(s.get("runs")),
        "SB":  fmt(s.get("stolenBases")),
        "OBP": fmt(s.get("obp"), 3),
        "SLG": fmt(s.get("slg"), 3),
        "OPS": fmt(s.get("ops"), 3),
    }

def format_pitching(s: dict) -> dict:
    def fmt(val, decimals=0):
        if val is None: return "—"
        try:
            v = float(val)
            if decimals: return f"{v:.{decimals}f}"
            return str(int(v))
        except: return str(val)

    return {
        "G":    fmt(s.get("gamesPlayed")),
        "GS":   fmt(s.get("gamesStarted")),
        "W":    fmt(s.get("wins")),
        "L":    fmt(s.get("losses")),
        "SV":   fmt(s.get("saves")),
        "IP":   fmt(s.get("inningsPitched"), 1),
        "ERA":  fmt(s.get("era"), 2),
        "WHIP": fmt(s.get("whip"), 2),
        "K":    fmt(s.get("strikeOuts")),
        "BB":   fmt(s.get("baseOnBalls")),
    }

# Pitchers by position tag
PITCHERS = {"RHP", "LHP", "SP", "RP", "P"}

def is_pitcher(position: str) -> bool:
    return any(p in position.upper() for p in PITCHERS)

# ── Game Log Fetching ─────────────────────────────────────────────────────────
def get_game_log(person_id: int, season: int, level: str = "", limit: int = 5) -> list[dict]:
    """Fetch recent game-by-game log for a player. Returns list of game dicts."""
    sport_id = LEVEL_SPORT_ID.get(level)
    if sport_id is None:
        return []
    pitcher = None  # determined from results
    games = []
    for group in ("pitching", "hitting"):
        try:
            r = requests.get(
                f"{MLB_API}/people/{person_id}/stats",
                params={"stats": "gameLog", "season": season,
                        "group": group, "sportId": sport_id,
                        "hydrate": "opponent"},
                headers=HEADERS, timeout=10
            )
            r.raise_for_status()
            splits = r.json().get("stats", [])
            if splits and splits[0].get("splits"):
                raw = splits[0]["splits"]
                for s in reversed(raw[-limit:]):  # most recent last → reverse to get newest first
                    g = s.get("stat", {})
                    game_info = s.get("game", {})
                    team_info = s.get("opponent", {})
                    raw_date  = s.get("date", "")[:10]
                    date_str  = (f"{raw_date[5:7]}-{raw_date[8:10]}-{raw_date[:4]}"
                                 if len(raw_date) == 10 else raw_date)
                    opp       = (team_info.get("abbreviation")
                                 or team_info.get("teamCode")
                                 or team_info.get("fileCode")
                                 or (team_info.get("name", "")[:3].upper() if team_info.get("name") else "???"))
                    is_home   = s.get("isHome", True)
                    opp_label = f"{'vs' if is_home else '@'} {opp}"
                    if group == "pitching":
                        games.append({
                            "date": date_str, "opp": opp_label, "group": "pitching",
                            "IP":   g.get("inningsPitched", "—"),
                            "H":    str(g.get("hits", "—")),
                            "ER":   str(g.get("earnedRuns", "—")),
                            "BB":   str(g.get("baseOnBalls", "—")),
                            "K":    str(g.get("strikeOuts", "—")),
                            "ERA":  g.get("era", "—"),
                        })
                    else:
                        ab  = g.get("atBats", 0) or 0
                        h   = g.get("hits",   0) or 0
                        hr  = g.get("homeRuns", 0) or 0
                        rbi = g.get("rbi",     0) or 0
                        bb  = g.get("baseOnBalls", 0) or 0
                        avg = f"{h/ab:.3f}".lstrip("0") if ab else "—"
                        games.append({
                            "date": date_str, "opp": opp_label, "group": "hitting",
                            "AB":  str(ab),
                            "H":   str(h),
                            "HR":  str(hr),
                            "RBI": str(rbi),
                            "BB":  str(bb),
                            "AVG": avg,
                        })
                if games:
                    break  # got data from this group, don't need the other
        except Exception as e:
            print(f"    ! Game log error for {person_id} ({group}): {e}")
    return games[:limit]


def format_game_log_html(games: list[dict], pitcher: bool) -> str:
    """Build an HTML table from game log entries."""
    if not games:
        return '<p style="color:#aaa;font-style:italic;font-size:0.85rem;margin-top:8px;">No recent game data available yet this season.</p>'
    if pitcher:
        headers = ["Date", "Opp", "IP", "H", "ER", "BB", "K"]
        keys    = ["date", "opp", "IP", "H", "ER", "BB", "K"]
    else:
        headers = ["Date", "Opp", "AB", "H", "HR", "RBI", "BB"]
        keys    = ["date", "opp", "AB", "H", "HR", "RBI", "BB"]
    th = "".join(f"<th>{h}</th>" for h in headers)
    rows = ""
    for g in games:
        hr_val = g.get("HR", "0")
        hr_cls = ' class="gl-hr"' if hr_val not in ("0", "—", "") else ""
        tds = ""
        for k in keys:
            val = g.get(k, "—")
            extra = hr_cls if k == "HR" and hr_val not in ("0","—","") else ""
            tds += f"<td{extra}>{val}</td>"
        rows += f"<tr>{tds}</tr>"
    return (f'<table class="modal-stats-table game-log-tbl">'
            f'<thead><tr>{th}</tr></thead>'
            f'<tbody>{rows}</tbody></table>')


# ── Update Excel Stats Sheet ──────────────────────────────────────────────────
def update_excel(player_data: list[dict]):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Remove old stats sheet if present
    if "2026 Stats" in wb.sheetnames:
        del wb["2026 Stats"]

    ws = wb.create_sheet("2026 Stats")

    header_fill  = PatternFill("solid", fgColor="FF1A3A5C")
    header_font  = Font(bold=True, color="FFFFFFFF", size=10)
    section_fill = PatternFill("solid", fgColor="FFD6E4F0")
    section_font = Font(bold=True, size=10)
    center       = Alignment(horizontal="center")
    left         = Alignment(horizontal="left")

    batter_headers  = ["Name","Pos","Team","Level","G","AB","AVG","HR","RBI","R","SB","OBP","SLG","OPS","Last Updated"]
    pitcher_headers = ["Name","Pos","Team","Level","G","GS","W","L","SV","IP","ERA","WHIP","K","BB","Last Updated"]

    updated = datetime.now().strftime("%m/%d/%Y %H:%M")
    row_num = 1

    def write_section(title, headers, players_subset):
        nonlocal row_num
        # Section header
        ws.cell(row_num, 1, title).font = section_font
        ws.cell(row_num, 1).fill = section_fill
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        row_num += 1

        # Column headers
        for col, h in enumerate(headers, 1):
            c = ws.cell(row_num, col, h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
        row_num += 1

        # Data rows
        for p in players_subset:
            stats = p.get("stats_fmt", {})
            row_vals = [p["name"], p["position"], p["team"], p["level"]]
            for h in headers[4:-1]:
                row_vals.append(stats.get(h, "—"))
            row_vals.append(updated)
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row_num, col, val)
                c.alignment = center if col > 2 else left
            row_num += 1
        row_num += 1  # blank row between sections

    # Split into pitchers and hitters
    levels_order = ["MLB", "AAA", "AA", "High-A", "Low-A", "Rookie", "Independent"]
    hitters  = [p for p in player_data if not is_pitcher(p["position"]) and p.get("stats_fmt")]
    pitchers = [p for p in player_data if is_pitcher(p["position"]) and p.get("stats_fmt")]

    # Sort each group by level
    level_rank = {l: i for i, l in enumerate(levels_order)}
    hitters.sort(key=lambda p: level_rank.get(p["level"], 99))
    pitchers.sort(key=lambda p: level_rank.get(p["level"], 99))

    write_section(f"⚾  HITTERS — {SEASON} Stats", batter_headers, hitters)
    write_section(f"🎯  PITCHERS — {SEASON} Stats", pitcher_headers, pitchers)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    for col in ["E","F","G","H","I","J","K","L","M","N"]:
        ws.column_dimensions[col].width = 8
    ws.column_dimensions["O"].width = 18

    wb.save(EXCEL_PATH)
    print(f"  ✓ Excel 'Stats' sheet updated ({len(player_data)} players)")

# ── Auto-News Generation ──────────────────────────────────────────────────────
LEVEL_RANK = {"MLB": 0, "AAA": 1, "AA": 2, "High-A": 3, "Low-A": 4, "Rookie": 5, "Independent": 6}

def load_news_cache() -> dict:
    if NEWS_CACHE_PATH.exists():
        with open(NEWS_CACHE_PATH) as f:
            return json.load(f)
    return {}

def save_news_cache(cache: dict):
    with open(NEWS_CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)

def _news_card(date_str: str, headline: str, body: str, tag: str) -> str:
    return (f'<div class="news-card">'
            f'<div class="news-date">{date_str}</div>'
            f'<h4>{headline}</h4>'
            f'<p>{body}</p>'
            f'<span class="news-tag">{tag}</span>'
            f'</div>')

def generate_news_cards(player_data: list[dict]) -> str:
    """Detect real events from live stats data and build news card HTML.

    Events detected (in priority order):
      1. MLB debut — player at MLB level who was not there last run
      2. Promotion — player moved up at least one level since last run
      3. Pitching milestone — ERA ≤ 2.00 with ≥ 5 IP
      4. Hitting milestone — AVG ≥ .340 with ≥ 10 AB, or ≥ 5 HR
      5. Season-summary fallback card (always included if < 4 events)
    """
    news_cache = load_news_cache()
    today = datetime.now()
    date_str = today.strftime("%B %Y")
    cards = []

    mlb_players  = [p for p in player_data if p.get("level") == "MLB"]
    mlb_count    = len(mlb_players)
    total        = len(player_data)
    orgs         = len(set(p["org"] for p in player_data if p.get("org")))

    # ── 0. Pinned call-up cards (manually confirmed, always first) ───────────
    pinned = [
        _news_card(
            "April 2026",
            "Jack Anderson Called Up to the Boston Red Sox",
            "The former FSU right-hander has earned his spot in The Show. "
            "Anderson was called up by Boston in April 2026 after working through "
            "the Red Sox minor league system. The pipeline delivers.",
            "Call-Up"
        ),
    ]
    cards.extend(pinned)  # always inserted, regardless of count
    MAX_CARDS = 4 + len(pinned)  # allow room for pinned + 4 auto cards

    # ── 1. MLB debuts ──────────────────────────────────────────────────────────
    for p in player_data:
        if len(cards) >= MAX_CARDS:
            break
        name  = p["name"]
        level = p.get("level", "")
        prev  = news_cache.get(name, {}).get("level", "")
        if level == "MLB" and prev and prev != "MLB":
            pos  = p.get("position", "")
            team = p.get("team", "")
            role = "left-hander" if "LHP" in pos else ("right-hander" if "RHP" in pos else "infielder" if pos in ("1B","2B","3B","SS") else "outfielder" if pos in ("LF","CF","RF","OF") else "catcher" if pos == "C" else "player")
            cards.append(_news_card(
                date_str,
                f"{name} Reaches the Major Leagues",
                f"Former Seminole {role} {name} has been promoted to the {team}, "
                f"earning a spot on an MLB roster this season. Another Nole makes it to The Show.",
                "MLB Debut"
            ))

    # ── 2. Promotions (non-MLB) ────────────────────────────────────────────────
    for p in player_data:
        if len(cards) >= MAX_CARDS:
            break
        name  = p["name"]
        level = p.get("level", "")
        prev  = news_cache.get(name, {}).get("level", "")
        if not prev or level == "MLB":
            continue
        if level in LEVEL_RANK and prev in LEVEL_RANK and LEVEL_RANK[level] < LEVEL_RANK[prev]:
            team = p.get("team", "")
            org  = p.get("org", "")
            org_note = f" in the {org} system" if org else ""
            cards.append(_news_card(
                date_str,
                f"{name} Promoted to {level}",
                f"FSU alum {name} has moved up to {level}{org_note}, now suiting up "
                f"for the {team}. The Seminole pipeline keeps producing.",
                "Promotion"
            ))

    # ── 3. Pitching milestones — ONE card, best ERA leads ─────────────────────
    pitching_candidates = []
    for p in player_data:
        if not is_pitcher(p.get("position", "")):
            continue
        stats = p.get("stats_fmt", {})
        try:
            era  = float(stats.get("ERA", "99").replace("—","99"))
            ip   = float(stats.get("IP",  "0").replace("—","0"))
            whip = float(stats.get("WHIP","99").replace("—","99"))
            ks   = int(stats.get("K", "0").replace("—","0"))
            gs   = int(stats.get("GS", "0").replace("—","0"))
        except ValueError:
            continue
        if ip >= 5 and era <= 3.50:
            pitching_candidates.append((era, ip, whip, ks, gs, p))
    pitching_candidates.sort(key=lambda x: x[0])  # best ERA first

    if pitching_candidates and len(cards) < MAX_CARDS:
        era, ip, whip, ks, gs, p = pitching_candidates[0]
        name  = p["name"]
        level = p.get("level", "")
        team  = p.get("team", "")
        pos   = p.get("position", "")
        hand  = "Left-hander" if "LHP" in pos else "Right-hander"
        role  = "starter" if gs >= 2 else "reliever"

        if era == 0.00:
            headline = f"{name} Hasn't Allowed an Earned Run All Season"
            body = (f"The FSU alum has been unhittable out of the {team} bullpen, throwing {ip:.1f} scoreless "
                    f"innings with {ks} strikeouts and a {whip:.2f} WHIP. "
                    f"A 0.00 ERA is as clean as it gets.")
        elif era < 1.00:
            headline = f"{name} Is Pitching to a {era:.2f} ERA for the {team.split()[-1]}"
            body = (f"The {hand.lower()} {role} has been dominant early at {level}, holding opponents to a "
                    f"{era:.2f} ERA across {ip:.1f} innings with {ks} strikeouts. "
                    f"His {whip:.2f} WHIP tells the same story.")
        elif era <= 2.00:
            headline = f"{name} Posting a {era:.2f} ERA Through {ip:.0f} Innings"
            body = (f"{hand} {name} has been tough to score on at {level} for the {team}, "
                    f"striking out {ks} hitters with a {whip:.2f} WHIP through {ip:.1f} innings. "
                    f"One of the better early-season lines among Seminole pitchers.")
        else:
            headline = f"{name} Solid Out of the {team.split()[-1]} {('Rotation' if role == 'starter' else 'Bullpen')}"
            body = (f"{hand} {name} has a {era:.2f} ERA through {ip:.1f} innings at {level}, "
                    f"with {ks} strikeouts and a {whip:.2f} WHIP for the {team}. "
                    f"A reliable arm in the Seminole pipeline.")
        cards.append(_news_card(date_str, headline, body, "Hot Start"))

    # ── 4. Hitting milestones — ONE card per distinct angle ────────────────────
    hitting_candidates = []
    for p in player_data:
        if is_pitcher(p.get("position", "")):
            continue
        stats = p.get("stats_fmt", {})
        try:
            avg = float(stats.get("AVG","0").replace("—","0") or "0")
            ab  = int(stats.get("AB", "0").replace("—","0"))
            hr  = int(stats.get("HR", "0").replace("—","0"))
            rbi = int(stats.get("RBI","0").replace("—","0"))
            ops = float(stats.get("OPS","0").replace("—","0") or "0")
            sb  = int(stats.get("SB", "0").replace("—","0"))
            h   = int(stats.get("H", "0").replace("—","0"))
        except ValueError:
            continue
        high_avg = ab >= 10 and avg >= 0.300
        power    = hr >= 3
        if high_avg or power:
            score = avg * 10 + hr * 0.5 + ops
            hitting_candidates.append((score, avg, ab, hr, rbi, ops, sb, h, p))
    hitting_candidates.sort(key=lambda x: -x[0])

    used_angles = set()
    for score, avg, ab, hr, rbi, ops, sb, h, p in hitting_candidates:
        if len(cards) >= MAX_CARDS:
            break
        name  = p["name"]
        level = p.get("level", "")
        team  = p.get("team", "")

        # Pick the most specific angle, skip if we've already run that angle
        if hr >= 3 and avg >= 0.300 and "five_tool" not in used_angles:
            angle = "five_tool"
            headline = f"{name} Hitting .{int(avg*1000):03d} with {hr} Home Runs for {team.split()[-1]}"
            body = (f"The former Seminole is doing everything right at {level} — solid contact and real power. "
                    f"His {rbi} RBI and {ops:.3f} OPS rank him among the most productive hitters in the league.")
        elif hr >= 3 and "power" not in used_angles:
            angle = "power"
            headline = f"{name} Already Has {hr} Home Runs at {level}"
            body = (f"The power is not a question for the FSU alum. {name} is slugging for the {team} with "
                    f"a {ops:.3f} OPS and {rbi} RBI. If the ball gets in the air, it's in trouble.")
        elif sb >= 3 and avg >= 0.280 and "speed" not in used_angles:
            angle = "speed"
            headline = f"{name} Batting .{int(avg*1000):03d} with {sb} Steals at {level}"
            body = (f"Speed and contact — the former Seminole is a handful at {level} for the {team}. "
                    f"{name} has {h} hits and {sb} stolen bases, making things happen every time he reaches.")
        elif "contact" not in used_angles:
            angle = "contact"
            # Highlight what's most eye-catching: the avg itself
            avg_display = f".{int(avg*1000):03d}"
            if avg >= 0.400:
                headline = f"{name} Is Hitting {avg_display} — Yes, Really"
                body = (f"The FSU alum is in a different stratosphere right now at {level}. "
                        f"{name} is batting {avg_display} for the {team} through {ab} at-bats, "
                        f"with {rbi} RBI and a {ops:.3f} OPS. Elite contact.")
            elif level == "MLB":
                headline = f"{name} Hitting {avg_display} to Open the MLB Season"
                body = (f"Not many Seminoles make it to The Show — and {name} is making the most of it. "
                        f"He's batting {avg_display} with {rbi} RBI for the {team}, posting a {ops:.3f} OPS through {ab} at-bats.")
            else:
                headline = f"{name} at .{int(avg*1000):03d} — Among the Best in {level}"
                body = (f"The former Seminole is one of the more consistent bats at {level} this season. "
                        f"{name} is hitting .{int(avg*1000):03d} with {rbi} RBI for the {team}, "
                        f"a {ops:.3f} OPS through {ab} at-bats.")
        else:
            continue  # already used all distinct angles

        used_angles.add(angle)
        cards.append(_news_card(date_str, headline, body, "Hot Bat"))

    # ── 5. Fallback season summary (fills remaining slots) ─────────────────────
    fallback_cards = [
        _news_card(
            date_str,
            "Stats Updating Daily",
            "Beyond Howser pulls live data from the MLB Stats API every morning. "
            "Roster news is updated manually as it happens. "
            "Bookmark this page and check back throughout the season for the latest numbers.",
            "Site Update"
        ),
    ]
    for fb in fallback_cards:
        if len(cards) >= MAX_CARDS:
            break
        cards.append(fb)

    # ── Update level cache for next run ───────────────────────────────────────
    for p in player_data:
        name  = p["name"]
        level = p.get("level", "")
        if level:
            entry = news_cache.get(name, {})
            entry["level"] = level
            entry["last_seen"] = today.strftime("%Y-%m-%d")
            news_cache[name] = entry
    save_news_cache(news_cache)

    return "\n    ".join(cards[:MAX_CARDS])


# ── Generate HTML Dashboard ───────────────────────────────────────────────────
def generate_html(player_data: list[dict], news_html: str = ""):
    updated = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    levels_order = ["MLB", "AAA", "AA", "High-A", "Low-A", "Rookie", "Independent", "60-Day IL", "7-Day IL"]
    special_levels = {"Released", "60-Day IL", "7-Day IL"}  # shown in separate section
    level_rank = {l: i for i, l in enumerate(levels_order)}
    # Separate active players from released
    active_players  = [p for p in player_data if p["level"] != "Released"]
    released_players = [p for p in player_data if p["level"] == "Released"]
    def sort_level(p):
        lvl = p["level"]
        if lvl in ("60-Day IL", "7-Day IL"):
            return level_rank.get(p.get("base_level") or lvl, 99)
        return level_rank.get(lvl, 99)
    sorted_players = sorted(active_players, key=sort_level)

    level_colors = {
        "MLB":         "#782F40",  # FSU Garnet
        "AAA":         "#B5451B",  # Burnt orange
        "AA":          "#CEB888",  # FSU Gold
        "High-A":      "#4A7C59",  # Forest green
        "Low-A":       "#2C5F8A",  # Steel blue
        "Rookie":      "#6B4C93",  # Purple
        "Independent": "#5a5a5a",  # Neutral gray
        "60-Day IL":   "#8B4513",  # Brown/rust
        "7-Day IL":    "#8B4513",  # Brown/rust
        "Released":    "#999999",  # Light gray
    }
    light_levels = {"AA"}  # light backgrounds need dark text

    # ── Dynamic hero stats ────────────────────────────────────────────────────
    total_players = len(active_players)
    mlb_count     = sum(1 for p in active_players if p.get("level") == "MLB")
    org_count     = len(set(p["org"] for p in active_players if p.get("org") and p.get("org") != "—"))

    # ── Photo URL helper ──────────────────────────────────────────────────────
    def photo_url(p):
        mid = p.get("mlb_id")
        if not mid:
            return ""
        base = "https://img.mlbstatic.com/mlb-photos/image/upload"
        # No Cloudinary fallback here — let onerror handle missing /67/ headshots
        # so MiLB-only players (like Drew Parrish) get their /milb/ photo instead
        return f"{base}/q_auto:best,f_auto,w_120/v1/people/{mid}/headshot/67/current"

    SHADOW_URL = "https://img.mlbstatic.com/mlb-photos/image/upload/d_people:generic:headshot:67:current.png,q_auto:best,f_auto,w_120/v1/people/1/headshot/67/current"
    PHOTO_ONERROR = (
        "var t=this,s=t.src;"
        "if(!t.dataset.tried1){"
        "t.dataset.tried1=1;"
        "t.src=s.replace('/67/','/milb/');"
        "}else if(!t.dataset.tried2){"
        "t.dataset.tried2=1;"
        f"t.src='{SHADOW_URL}';"
        "t.onerror=null;"
        "}"
    )

    # ── Card view ─────────────────────────────────────────────────────────────
    modal_data = []

    def player_card(p, idx):
        stats    = p.get("stats_fmt", {})
        lvl      = p["level"]
        base_lvl = p.get("base_level", "") or lvl
        badge_lvl = base_lvl if lvl in ("60-Day IL", "7-Day IL") else lvl
        color    = level_colors.get(badge_lvl, "#555")
        txt      = "#333" if badge_lvl in light_levels else "white"
        pitcher = is_pitcher(p["position"])
        stat_keys = ([("ERA","ERA"),("IP","IP"),("W","W"),("L","L"),
                      ("SV","SV"),("K","K"),("BB","BB"),("WHIP","WHIP")]
                     if pitcher else
                     [("AVG","AVG"),("HR","HR"),("RBI","RBI"),("OPS","OPS"),("OBP","OBP"),
                      ("G","G"),("AB","AB"),("H","H"),("R","R"),("SB","SB")])
        stats_html = "".join(
            f'<div class="stat"><div class="stat-val">{stats.get(k,"—")}</div>'
            f'<div class="stat-lbl">{lbl}</div></div>'
            for k, lbl in stat_keys
        )
        no_data_msg = ('<div class="no-data">Season not started or not in MLB system</div>'
                       if not stats else "")
        org_line = (f'<div class="card-org">⬆ {p["org"]}</div>'
                    if lvl != "MLB" and p.get("org") else "")
        purl = photo_url(p)
        photo_tag = (f'<img class="card-photo" src="{purl}" alt="{p["name"]}" onerror="{PHOTO_ONERROR}">'
                     if purl else
                     f'<img class="card-photo" src="{SHADOW_URL}" alt="{p["name"]}">')

        # Build modal stats table for this player
        if pitcher:
            mhdrs = ["G","GS","W","L","SV","IP","ERA","WHIP","K","BB"]
        else:
            mhdrs = ["G","AB","AVG","HR","RBI","R","SB","OBP","SLG","OPS"]
        th = "".join(f"<th>{h}</th>" for h in mhdrs)
        td = "".join(f"<td>{stats.get(h,'—')}</td>" for h in mhdrs)
        stats_tbl = (f'<table class="modal-stats-table"><thead><tr>{th}</tr></thead>'
                     f'<tbody><tr>{td}</tr></tbody></table>' if stats else
                     '<p style="color:#aaa;font-style:italic;font-size:0.85rem;">No stats available yet this season.</p>')

        initials = "".join(w[0] for w in p["name"].split()[:2]).upper()
        game_log_html = format_game_log_html(p.get("game_log", []), pitcher)
        modal_data.append({
            "name":        p["name"],
            "posTeam":     f'{p["position"]} · {p["team"]}',
            "lvl":         lvl,
            "color":       color,
            "txt":         txt,
            "draft":       p.get("draft", ""),
            "notes":       p.get("notes", ""),
            "photo":       purl,
            "initials":    initials,
            "statsHtml":   stats_tbl,
            "gameLogHtml": game_log_html,
            "milbUrl":     p.get("milb_url", ""),
        })

        il_badge = (f'<span style="display:inline-block;background:#8B4513;color:white;'
                    f'font-size:0.65rem;font-weight:700;padding:2px 7px;border-radius:10px;'
                    f'margin-left:6px;vertical-align:middle;">{lvl}</span>'
                    if lvl in ("60-Day IL", "7-Day IL") else "")

        return f'''
        <div class="card" data-level="{lvl}" data-name="{p["name"].lower()}" onclick="openModal({idx})" style="cursor:pointer">
          <div class="card-header" style="background:{color};color:{txt}">
            {photo_tag}
            <div class="card-info">
              <div class="card-name">{p["name"]}{il_badge}</div>
              <div class="card-meta">{p["position"]} · {p["team"]}</div>
              {org_line}
            </div>
          </div>
          <div class="card-level" style="color:{color}">{badge_lvl}</div>
          <div class="card-stats">{stats_html}{no_data_msg}</div>
        </div>'''

    cards_html = "\n".join(player_card(p, i) for i, p in enumerate(sorted_players))

    # ── List/table view ───────────────────────────────────────────────────────
    def table_row(p):
        lvl        = p["level"]
        base_lvl   = p.get("base_level", "") or lvl  # use base level for IL players' badge color
        badge_lvl  = base_lvl if lvl in ("60-Day IL", "7-Day IL") else lvl
        color      = level_colors.get(badge_lvl, "#555")
        txt        = "#333" if badge_lvl in light_levels else "white"
        stats      = p.get("stats_fmt", {})
        pitcher    = is_pitcher(p["position"])
        if pitcher:
            s1_lbl, s1_val = "ERA",  stats.get("ERA", "—")
            s2_lbl, s2_val = "IP",   stats.get("IP",  "—")
            s3_lbl, s3_val = "WHIP", stats.get("WHIP","—")
        else:
            s1_lbl, s1_val = "AVG",  stats.get("AVG", "—")
            s2_lbl, s2_val = "HR",   stats.get("HR",  "—")
            s3_lbl, s3_val = "OPS",  stats.get("OPS", "—")
        org_cell = p.get("org", "") if lvl != "MLB" else "—"
        il_tag = (f'<span style="display:inline-block;background:#8B4513;color:white;'
                  f'font-size:0.6rem;font-weight:700;padding:1px 6px;border-radius:8px;'
                  f'margin-left:5px;vertical-align:middle;">{lvl}</span>'
                  if lvl in ("60-Day IL", "7-Day IL") else "")
        name_link = f'<a href="{p["milb_url"]}" target="_blank" rel="noopener" style="color:inherit;text-decoration:none;">{p["name"]}</a>' if p.get("milb_url") else p["name"]
        return f'''
        <tr data-level="{lvl}" data-name="{p["name"].lower()}">
          <td class="td-name">
            <img class="row-photo" src="{photo_url(p) or SHADOW_URL}" alt="{p["name"]}" onerror="{PHOTO_ONERROR}">
            <span class="name-text">{name_link}{il_tag} <span class="pos-tag">{p["position"]}</span></span>
          </td>
          <td><span class="lvl-badge" style="background:{color};color:{txt}">{badge_lvl}</span></td>
          <td class="td-team">{p["team"]}</td>
          <td class="td-org">{org_cell}</td>
          <td class="td-stat"><span class="stat-lbl-sm">{s1_lbl}</span> {s1_val}</td>
          <td class="td-stat"><span class="stat-lbl-sm">{s2_lbl}</span> {s2_val}</td>
          <td class="td-stat"><span class="stat-lbl-sm">{s3_lbl}</span> {s3_val}</td>
        </tr>'''

    rows_html = "\n".join(table_row(p) for p in sorted_players)

    # ── Released players section ───────────────────────────────────────────────
    def released_card(p):
        return f'''
        <div class="released-card">
          <div class="released-name">{p["name"]}</div>
          <div class="released-pos">{p["position"]}</div>
          <div class="released-status">Released</div>
          <div class="released-note">{p.get("draft","")}</div>
        </div>'''

    released_html = ""
    if released_players:
        released_cards = "\n".join(released_card(p) for p in released_players)
        released_html = f'''
<div class="section-wrap" id="released" style="margin-top:0;padding-top:0;">
  <div class="section-title" style="color:#999;font-size:1rem;">No Longer Active</div>
  <div class="released-grid">
    {released_cards}
  </div>
</div>'''

    # ── Level filter buttons (color-coordinated) ──────────────────────────────
    level_btns = ('<button class="filter-btn active" data-color="#782F40" '
                  'onclick="filterLevel(\'all\',this)">All</button>\n')
    for lvl in levels_order:
        if any(p["level"] == lvl for p in player_data):
            c = level_colors.get(lvl, "#555")
            level_btns += (f'<button class="filter-btn" data-color="{c}" '
                           f'onclick="filterLevel(\'{lvl}\',this)">{lvl}</button>\n')

    modal_json = json.dumps(modal_data, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Beyond Howser — {SEASON} Stats</title>
<link rel="icon" type="image/png" href="logo.png">
<link rel="shortcut icon" type="image/png" href="logo.png">
<script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client=ca-pub-4259331118703482" crossorigin="anonymous"></script>
<style>
:root {{
  --garnet: #782F40; --garnet-dark: #5a1f2d; --garnet-light: #9e4055;
  --gold: #CEB888;   --gold-light: #e8d9b0;  --gold-dark: #a8955e;
  --cream: #FAF7F2;  --border: #e0d8d0;
  --shadow: 0 2px 12px rgba(120,47,64,0.10);
}}
*,*::before,*::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
html {{ scroll-behavior: smooth; }}
body {{ font-family: 'Segoe UI', system-ui, sans-serif; background: var(--cream); color: #1a1a1a; line-height: 1.5; }}

/* ── Nav ── */
nav {{ background: var(--garnet); padding: 0 32px; display: flex; align-items: center;
       justify-content: space-between; position: sticky; top: 0; z-index: 100;
       box-shadow: 0 2px 8px rgba(0,0,0,0.25); }}
.nav-brand {{ display: flex; align-items: center; gap: 10px; padding: 14px 0; text-decoration: none; }}
.nav-logo {{ width: 42px; height: 42px; border-radius: 6px; flex-shrink: 0;
              object-fit: cover; display: block; }}
.nav-title {{ color: white; font-weight: 700; font-size: 1.1rem; }}
.nav-sub   {{ color: var(--gold); font-size: 0.7rem; letter-spacing: 0.08em; text-transform: uppercase; }}
.nav-links {{ display: flex; gap: 2px; margin: 0 24px; }}
.nav-links a {{ color: rgba(255,255,255,0.8); text-decoration: none; padding: 8px 14px;
                border-radius: 6px; font-size: 0.88rem; transition: all 0.15s; white-space: nowrap; }}
.nav-links a:hover {{ background: rgba(255,255,255,0.12); color: white; }}
.nav-links a.active {{ background: rgba(206,184,136,0.2); color: var(--gold); font-weight: 600; }}
.nav-updated {{ color: rgba(255,255,255,0.70); font-size: 0.72rem; margin-left: auto; white-space: nowrap; }}

/* ── Hero ── */
.hero {{ background: linear-gradient(135deg, var(--garnet-dark) 0%, var(--garnet) 55%, var(--garnet-light) 100%);
         padding: 28px 32px 24px; text-align: center; position: relative; overflow: hidden; }}
.hero::before {{ content:''; position: absolute; inset: 0;
  background: repeating-linear-gradient(45deg, transparent, transparent 40px,
    rgba(206,184,136,0.04) 40px, rgba(206,184,136,0.04) 80px); }}
.hero-content {{ position: relative; z-index: 1; max-width: 700px; margin: 0 auto; }}
.hero-eyebrow {{ display: inline-block; background: rgba(206,184,136,0.2);
  border: 1px solid rgba(206,184,136,0.4); color: var(--gold);
  font-size: 0.72rem; letter-spacing: 0.12em; text-transform: uppercase;
  padding: 4px 14px; border-radius: 20px; margin-bottom: 16px; }}
.hero h1 {{ color: white; font-size: 1.9rem; font-weight: 800; line-height: 1.15; margin-bottom: 18px; }}
.hero h1 span {{ color: var(--gold); }}
.hero-stats {{ display: flex; justify-content: center; gap: 40px;
               border-top: 1px solid rgba(206,184,136,0.2); padding-top: 18px; }}
.hero-stat .num {{ color: var(--gold); font-size: 2rem; font-weight: 800; line-height: 1; }}
.hero-stat .lbl {{ color: rgba(255,255,255,0.6); font-size: 0.72rem; text-transform: uppercase;
                    letter-spacing: 0.08em; margin-top: 4px; }}
.hero-updated {{ margin-top: 16px; padding-top: 14px; border-top: 1px solid rgba(206,184,136,0.2);
                  color: rgba(255,255,255,0.65); font-size: 0.8rem; letter-spacing: 0.03em; }}
.hero-updated strong {{ color: var(--gold); font-weight: 600; }}

/* ── Legacy accolades strip ── */
.legacy {{ background: var(--garnet-dark); border-bottom: 3px solid var(--gold); }}
.legacy-inner {{ max-width: 1100px; margin: 0 auto; padding: 20px 32px;
                 display: flex; gap: 0; flex-wrap: wrap; justify-content: space-around; }}
.legacy-item {{ text-align: center; padding: 10px 24px; border-right: 1px solid rgba(206,184,136,0.2); }}
.legacy-item:last-child {{ border-right: none; }}
.legacy-title {{ color: rgba(255,255,255,0.65); font-size: 0.68rem; font-weight: 600;
                 text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 5px; }}
.legacy-val {{ color: var(--gold); font-size: 1.25rem; font-weight: 800; line-height: 1.2; }}

/* ── Sections ── */
.section-wrap {{ max-width: 1200px; margin: 48px auto 0; padding: 0 32px; }}
.section-title {{ font-size: 1.3rem; font-weight: 700; color: var(--garnet);
                  border-left: 4px solid var(--gold); padding-left: 12px; margin-bottom: 20px; }}
.news-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(260px, 1fr)); gap: 16px; }}
.news-card {{ background: white; border: 1px solid var(--border); border-radius: 10px;
              padding: 18px 20px; box-shadow: var(--shadow); }}
.news-date {{ font-size: 0.72rem; color: #999; text-transform: uppercase;
              letter-spacing: 0.06em; margin-bottom: 7px; }}
.news-card h4 {{ font-size: 0.95rem; font-weight: 700; color: var(--garnet);
                  margin-bottom: 7px; line-height: 1.35; }}
.news-card p {{ font-size: 0.83rem; color: #666; line-height: 1.55; }}
.news-tag {{ display: inline-block; background: var(--gold-light); color: var(--gold-dark);
             font-size: 0.68rem; font-weight: 700; text-transform: uppercase;
             letter-spacing: 0.06em; padding: 2px 9px; border-radius: 10px; margin-top: 10px; }}
.about-box {{ background: white; border: 1px solid var(--border); border-radius: 10px;
              padding: 28px 32px; box-shadow: var(--shadow); line-height: 1.7; color: #444; }}
.about-box p {{ margin-bottom: 14px; font-size: 0.92rem; }}
.about-box p:last-child {{ margin-bottom: 0; }}
.about-box strong {{ color: var(--garnet); }}
.news-grid-wrap {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(240px,1fr)); gap: 16px; }}
.released-grid {{ display: flex; flex-wrap: wrap; gap: 12px; padding: 4px 0 16px; }}
.released-card {{ background: white; border: 1px solid #ddd; border-radius: 8px; padding: 12px 16px;
                   display: flex; align-items: center; gap: 14px; min-width: 220px; }}
.released-name {{ font-weight: 700; font-size: 0.9rem; color: #555; }}
.released-pos {{ font-size: 0.78rem; color: #999; }}
.released-status {{ background: #999; color: white; font-size: 0.65rem; font-weight: 700;
                    padding: 2px 8px; border-radius: 10px; white-space: nowrap; }}
.released-note {{ font-size: 0.72rem; color: #bbb; display: none; }}

/* ── Controls bar ── */
.controls {{ background: white; padding: 14px 32px; border-bottom: 1px solid var(--border);
             display: flex; gap: 8px; flex-wrap: wrap; align-items: center;
             box-shadow: 0 1px 4px rgba(0,0,0,0.05); position: sticky; top: 64px; z-index: 90; }}
.controls-label {{ font-size: 0.78rem; font-weight: 700; color: #888;
                    text-transform: uppercase; letter-spacing: 0.05em; margin-right: 4px; }}
.filter-btn {{ padding: 5px 13px; border: 1.5px solid #ccc; background: white;
               border-radius: 20px; cursor: pointer; font-size: 0.78rem;
               color: #555; transition: all .15s; font-weight: 500; }}
.filter-btn:hover {{ border-color: var(--garnet); color: var(--garnet); }}
.search-box {{ padding: 6px 12px; border: 1.5px solid #ccc; border-radius: 20px;
               font-size: 0.82rem; width: 190px; outline: none; transition: border-color .15s; }}
.search-box:focus {{ border-color: var(--garnet); }}
.view-toggle {{ margin-left: auto; display: flex; gap: 4px; }}
.view-btn {{ padding: 5px 14px; border: 1.5px solid #ccc; background: white;
             border-radius: 20px; cursor: pointer; font-size: 0.78rem;
             color: #555; transition: all .15s; font-weight: 600; }}
.view-btn.active {{ background: var(--garnet); color: white; border-color: var(--garnet); }}
.controls-updated {{ margin-left: 12px; font-size: 0.72rem; color: #999; white-space: nowrap;
                      padding: 4px 0; display: flex; align-items: center; gap: 4px; }}
.controls-updated::before {{ content: '↻'; font-size: 0.78rem; color: var(--garnet); }}

/* ── Two-column roster layout ── */
.roster-layout {{ display: flex; gap: 24px; align-items: flex-start;
                  max-width: 1400px; margin: 0 auto; padding: 24px 32px; }}
.roster-main {{ flex: 1; min-width: 0; }}
.roster-sidebar {{ width: 300px; flex-shrink: 0; position: sticky; top: 118px;
                   display: flex; flex-direction: column; gap: 20px; }}

/* ── Sidebar widgets ── */
.sidebar-widget {{ background: white; border: 1px solid var(--border); border-radius: 10px;
                   overflow: hidden; box-shadow: var(--shadow); }}
.sidebar-widget-hdr {{ background: var(--garnet); color: white; padding: 11px 16px;
                        display: flex; align-items: center; gap: 8px;
                        font-weight: 700; font-size: 0.85rem; }}
.sidebar-widget-hdr a {{ color: var(--gold); text-decoration: none; font-size: 0.75rem;
                          margin-left: auto; border: 1px solid var(--gold);
                          padding: 2px 9px; border-radius: 12px; white-space: nowrap; }}
.sidebar-widget-hdr a:hover {{ background: var(--gold); color: var(--garnet); }}
.links-list {{ list-style: none; padding: 8px 0; }}
.links-list li {{ border-bottom: 1px solid var(--border); }}
.links-list li:last-child {{ border-bottom: none; }}
.links-list a {{ display: flex; align-items: center; gap: 10px; padding: 10px 16px;
                 color: #444; text-decoration: none; font-size: 0.85rem;
                 transition: background .12s; }}
.links-list a:hover {{ background: #fff5f6; color: var(--garnet); }}
.links-list .link-icon {{ width: 36px; height: 36px; border-radius: 8px; flex-shrink: 0;
  display: flex; align-items: center; justify-content: center; }}
.links-list .link-icon svg {{ width: 20px; height: 20px; }}
.links-list .link-lbl {{ font-weight: 600; }}
.links-list .link-sub {{ font-size: 0.72rem; color: #aaa; display: block; margin-top: 1px; }}

/* ── Card grid ── */
.grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 16px; }}
.card {{ background: white; border-radius: 10px; overflow: hidden;
         box-shadow: var(--shadow); transition: transform .15s, box-shadow .15s; }}
.card:hover {{ transform: translateY(-2px); box-shadow: 0 6px 20px rgba(120,47,64,0.15); }}
.card.hidden {{ display: none; }}
.card-header {{ padding: 14px 16px; color: white; display: flex; align-items: center; gap: 12px; }}
.card-photo {{ width: 54px; height: 54px; border-radius: 50%; object-fit: cover;
               border: 2px solid rgba(255,255,255,0.4); flex-shrink: 0; background: rgba(255,255,255,0.15); }}
.card-info {{ flex: 1; min-width: 0; }}
.card-name {{ font-size: 1rem; font-weight: 700; }}
.card-meta {{ font-size: 0.75rem; opacity: 0.85; margin-top: 2px; }}
.card-org  {{ font-size: 0.65rem; opacity: 0.75; margin-top: 3px; font-style: italic; }}
.card-level {{ font-size: 0.68rem; font-weight: 700; letter-spacing: .5px;
               text-transform: uppercase; padding: 6px 16px 0; }}
.card-stats {{ display: grid; grid-template-columns: repeat(5, 1fr); padding: 10px 12px 14px; gap: 6px; }}
.stat {{ text-align: center; }}
.stat-val {{ font-size: 1rem; font-weight: 700; color: var(--garnet); }}
.stat-lbl {{ font-size: 0.62rem; color: #999; text-transform: uppercase; margin-top: 1px; }}
.no-data {{ grid-column: 1/-1; text-align: center; color: #bbb;
            font-size: 0.78rem; padding: 8px 0; font-style: italic; }}
.season-note {{ text-align: center; font-size: 0.68rem; color: #aaa;
                font-style: italic; padding: 2px 0 6px; }}

/* ── List/Table view ── */
.list-wrap {{ display: none; }}
.roster-card {{ background: white; border-radius: 10px; border: 1px solid var(--border);
                overflow: hidden; box-shadow: var(--shadow); }}
table {{ width: 100%; border-collapse: collapse; font-size: 0.88rem; }}
thead tr {{ background: var(--garnet); }}
thead th {{ color: white; font-size: 0.72rem; font-weight: 600; letter-spacing: 0.06em;
            text-transform: uppercase; padding: 12px 14px; text-align: left; white-space: nowrap; }}
tbody tr {{ border-bottom: 1px solid var(--border); transition: background .1s; }}
tbody tr:last-child {{ border-bottom: none; }}
tbody tr:hover {{ background: #fff5f6; }}
tbody tr.hidden {{ display: none; }}
td {{ padding: 10px 14px; vertical-align: middle; }}
.td-name {{ display: flex; align-items: center; gap: 10px; font-weight: 600; white-space: nowrap; }}
.row-photo {{ width: 38px; height: 38px; border-radius: 50%; object-fit: cover;
              border: 2px solid var(--border); flex-shrink: 0; background: var(--gold-light); }}
.pos-tag {{ font-size: 0.72rem; font-weight: 400; color: #888; margin-left: 4px; }}
.lvl-badge {{ display: inline-block; padding: 3px 8px; border-radius: 12px;
              font-size: 0.7rem; font-weight: 700; letter-spacing: 0.04em;
              text-transform: uppercase; white-space: nowrap; }}
.td-team {{ color: #444; font-size: 0.85rem; }}
.td-org  {{ color: #777; font-size: 0.82rem; }}
.td-stat {{ font-size: 0.85rem; font-weight: 600; color: var(--garnet); white-space: nowrap; }}
.stat-lbl-sm {{ font-size: 0.68rem; font-weight: 400; color: #aaa; text-transform: uppercase; margin-right: 2px; }}

/* ── Footer ── */
footer {{ background: var(--garnet-dark); color: rgba(255,255,255,0.55);
          text-align: center; padding: 22px 32px; font-size: 0.8rem; margin-top: 48px; }}
footer a {{ color: var(--gold); text-decoration: none; }}

/* ── Responsive: Tablet (≤900px) ── */
@media (max-width: 900px) {{
  .roster-layout {{ flex-direction: column; padding: 16px 20px; }}
  .roster-sidebar {{ width: 100%; position: static; flex-direction: row; flex-wrap: wrap; gap: 16px; }}
  .sidebar-widget {{ flex: 1; min-width: 260px; }}
  .legacy-inner {{ gap: 0; }}
  .legacy-item {{ padding: 10px 14px; }}
  .section-wrap {{ padding: 0 20px; }}
  .controls {{ padding: 12px 20px; top: 58px; }}
}}

/* ── Responsive: Mobile (≤600px) ── */
@media (max-width: 600px) {{
  nav {{ padding: 0 16px; }}
  .nav-links {{ display: none; }}
  .nav-updated {{ display: none; }}
  .hero {{ padding: 20px 16px 18px; }}
  .hero h1 {{ font-size: 1.4rem; }}
  .hero-eyebrow {{ font-size: 0.65rem; }}
  .hero-stats {{ gap: 16px; flex-wrap: wrap; }}
  .hero-stat .num {{ font-size: 1.5rem; }}
  .legacy-inner {{ padding: 12px 16px; justify-content: flex-start; }}
  .legacy-item {{ padding: 8px 12px; width: 50%; border-right: none;
                  border-bottom: 1px solid rgba(206,184,136,0.2); }}
  .legacy-item:nth-child(odd) {{ border-right: 1px solid rgba(206,184,136,0.2); }}
  .legacy-val {{ font-size: 1rem; }}
  .controls {{ padding: 10px 16px; gap: 6px; top: 54px; }}
  .controls-label {{ display: none; }}
  .search-box {{ width: 130px; }}
  .view-toggle {{ margin-left: 0; }}
  .roster-layout {{ padding: 12px 8px; }}
  .grid {{ grid-template-columns: repeat(auto-fill, minmax(160px, 1fr)); gap: 10px; }}
  .card-stats {{ grid-template-columns: repeat(5, 1fr); gap: 4px; padding: 8px 8px 10px; }}
  .stat-val {{ font-size: 0.85rem; }}
  .stat-lbl {{ font-size: 0.55rem; }}
  .card-photo {{ width: 42px; height: 42px; }}
  .card-name {{ font-size: 0.88rem; }}
  .card-meta {{ font-size: 0.68rem; }}
  .roster-sidebar {{ flex-direction: column; }}
  .sidebar-widget {{ min-width: unset; }}
  .section-wrap {{ padding: 0 16px; }}
  .news-grid-wrap {{ grid-template-columns: 1fr; }}
  .about-box {{ padding: 18px 16px; }}
  table {{ font-size: 0.78rem; }}
  td, thead th {{ padding: 8px 10px; }}
  .row-photo {{ width: 28px; height: 28px; }}
  footer {{ padding: 18px 16px; }}
}}

/* ── Modal ── */
.modal-overlay {{ display:none; position:fixed; inset:0; background:rgba(0,0,0,.55);
  z-index:1000; align-items:center; justify-content:center; padding:20px; }}
.modal-overlay.open {{ display:flex; }}
.modal-box {{ background:white; border-radius:14px; width:100%; max-width:520px;
  box-shadow:0 20px 60px rgba(0,0,0,.3); animation:modalIn .2s ease; overflow:hidden; }}
@keyframes modalIn {{ from {{opacity:0;transform:translateY(-20px)}} to {{opacity:1;transform:translateY(0)}} }}
.modal-header {{ padding:18px 20px; position:relative; }}
.modal-header h2 {{ margin:0; font-size:1.2rem; font-weight:700; }}
.modal-header .mh-sub {{ font-size:0.78rem; opacity:.8; margin-top:4px; }}
.modal-close {{ position:absolute; top:14px; right:16px; background:rgba(255,255,255,.2);
  border:none; color:white; width:28px; height:28px; border-radius:50%; cursor:pointer;
  font-size:1.1rem; display:flex; align-items:center; justify-content:center; line-height:1; }}
.modal-close:hover {{ background:rgba(255,255,255,.35); }}
.modal-body {{ padding:20px; }}
.modal-photo-wrap {{ position:relative; width:72px; height:72px; float:right;
  margin:0 0 12px 16px; flex-shrink:0; }}
.modal-photo-init {{ position:absolute; inset:0; border-radius:50%; background:var(--garnet);
  color:white; display:flex; align-items:center; justify-content:center;
  font-weight:700; font-size:1.1rem; }}
.modal-photo {{ position:absolute; inset:0; width:100%; height:100%;
  border-radius:50%; object-fit:cover; }}
.modal-meta {{ font-size:0.85rem; color:#555; line-height:1.8; margin-bottom:12px; }}
.modal-meta strong {{ color:#222; }}
.modal-draft {{ font-size:0.8rem; color:#888; margin-bottom:6px; }}
.modal-notes {{ background:var(--gold-light); border-left:3px solid var(--gold-dark);
  padding:8px 12px; border-radius:0 6px 6px 0; font-size:0.82rem;
  color:#555; margin-bottom:14px; font-style:italic; }}
.modal-tabs {{ display:flex; gap:6px; margin:12px 0 8px; border-bottom:2px solid var(--border); }}
.modal-tab {{ background:none; border:none; border-bottom:3px solid transparent; padding:6px 14px;
              font-size:0.82rem; font-weight:600; color:#888; cursor:pointer; margin-bottom:-2px;
              transition:color 0.15s,border-color 0.15s; }}
.modal-tab.active {{ color:var(--garnet); border-bottom-color:var(--garnet); }}
.modal-tab:hover {{ color:var(--garnet); }}
.game-log-tbl {{ margin-top:4px; }}
.gl-hr {{ color:#c0392b; font-weight:700; }}
.modal-stats-table {{ width:100%; border-collapse:collapse; font-size:0.82rem; margin-top:4px; }}
.modal-stats-table th {{ background:var(--garnet); color:white; padding:6px 8px;
  text-align:center; font-size:0.72rem; letter-spacing:.04em; }}
.modal-stats-table td {{ padding:7px 8px; border-bottom:1px solid #eee;
  text-align:center; font-weight:600; color:var(--garnet); }}
.modal-milb-link {{ display:inline-block; margin-top:12px; font-size:0.8rem;
  color:var(--garnet); border:1px solid var(--garnet); padding:4px 12px;
  border-radius:12px; text-decoration:none; }}
.modal-milb-link:hover {{ background:var(--garnet); color:white; }}
.card-photo-init {{ width:54px; height:54px; border-radius:50%; background:rgba(255,255,255,.2);
  display:flex; align-items:center; justify-content:center; font-weight:700;
  font-size:1rem; flex-shrink:0; }}
</style>
</head>
<body>

<!-- Modal overlay -->
<div class="modal-overlay" id="modalOverlay" onclick="maybeClose(event)">
  <div class="modal-box">
    <div class="modal-header" id="mHeader">
      <h2 id="mName"></h2>
      <div class="mh-sub" id="mSub"></div>
      <button class="modal-close" onclick="closeModal()">✕</button>
    </div>
    <div class="modal-body">
      <div class="modal-photo-wrap">
        <div class="modal-photo-init" id="mPhotoInit"></div>
        <img class="modal-photo" id="mPhoto" src="" alt="">
      </div>
      <div class="modal-draft" id="mDraft"></div>
      <div class="modal-notes" id="mNotes" style="display:none"></div>
      <div class="modal-tabs">
        <button class="modal-tab active" id="tabStats" onclick="switchTab('stats')">Season Stats</button>
        <button class="modal-tab" id="tabLog"   onclick="switchTab('log')">Recent Games</button>
      </div>
      <div id="mStatsSection"></div>
      <div id="mGameLogSection" style="display:none"></div>
      <div style="clear:both"></div>
      <a class="modal-milb-link" id="mMilbLink" href="#" target="_blank" rel="noopener" style="display:none">View on MiLB.com ↗</a>
    </div>
  </div>
</div>

<!-- Nav -->
<nav>
  <a href="#home" class="nav-brand">
    <img class="nav-logo" src="logo.png" alt="Beyond Howser">
    <div>
      <div class="nav-title">Beyond Howser</div>
      <div class="nav-sub">FSU Baseball Alumni Tracker</div>
    </div>
  </a>
  <div class="nav-links">
    <a href="#home"   class="nav-link active">Home</a>
    <a href="#roster" class="nav-link">Roster</a>
    <a href="#news"   class="nav-link">News</a>
    <a href="#about"  class="nav-link">About</a>
  </div>
  <span class="nav-updated">Updated {updated}</span>
</nav>

<!-- Hero -->
<section class="hero" id="home">
  <div class="hero-content">
    <div class="hero-eyebrow">Florida State University</div>
    <h1>Tracking Every <span>Seminole</span> in Pro Baseball</h1>
    <div class="hero-stats">
      <div class="hero-stat"><div class="num">{total_players}</div><div class="lbl">Total Players</div></div>
      <div class="hero-stat"><div class="num">{mlb_count}</div><div class="lbl">On MLB Rosters</div></div>
      <div class="hero-stat"><div class="num">{org_count}</div><div class="lbl">Organizations</div></div>
      <div class="hero-stat"><div class="num">{SEASON}</div><div class="lbl">Season</div></div>
    </div>
    <div class="hero-updated">📅 Stats last updated: <strong>{updated} ET</strong></div>
  </div>
</section>

<!-- FSU Legacy Accolades -->
<div class="legacy">
  <div class="legacy-inner">
    <div class="legacy-item">
      <div class="legacy-title">Location</div>
      <div class="legacy-val">Tallahassee, FL</div>
    </div>
    <div class="legacy-item">
      <div class="legacy-title">Home Stadium</div>
      <div class="legacy-val">Dick Howser Stadium</div>
    </div>
    <div class="legacy-item">
      <div class="legacy-title">Conference</div>
      <div class="legacy-val">Atlantic Coast Conference</div>
    </div>
    <div class="legacy-item">
      <div class="legacy-title">CWS Appearances</div>
      <div class="legacy-val">24</div>
    </div>
  </div>
</div>

<!-- News Section (above roster) -->
<div class="section-wrap" id="news">
  <div class="section-title">Latest News &amp; Updates</div>
  <div class="news-grid-wrap">
    {news_html}
  </div>
</div>

<!-- Controls (sticky) -->
<div class="controls" id="roster">
  <span class="controls-label">Level:</span>
  {level_btns}
  <input class="search-box" type="text" placeholder="Search player…" oninput="applySearch(this.value)">
  <span class="controls-updated">Updated {updated} ET</span>
  <div class="view-toggle">
    <button class="view-btn active" id="btnCards" onclick="setView('cards')">⊞ Cards</button>
    <button class="view-btn" id="btnList"  onclick="setView('list')">≡ List</button>
  </div>
</div>

<!-- Two-column roster layout -->
<div class="roster-layout">

  <!-- Main: cards + list -->
  <div class="roster-main">
    <div class="grid" id="grid">
{cards_html}
    </div>
    <div class="list-wrap" id="listWrap">
      <div class="roster-card">
        <table>
          <thead>
            <tr>
              <th>Player</th><th>Level</th><th>Team</th><th>Organization</th><th colspan="3">Key Stats</th>
            </tr>
          </thead>
          <tbody id="listBody">
{rows_html}
          </tbody>
        </table>
      </div>
    </div>

{released_html}

  </div>

  <!-- Sidebar -->
  <div class="roster-sidebar">

    <!-- Follow us -->
    <div class="sidebar-widget">
      <div class="sidebar-widget-hdr">
        <span>Follow the Noles</span>
      </div>
      <div style="padding:16px 14px;display:flex;flex-direction:column;gap:10px;">
        <p style="font-size:0.82rem;color:#555;line-height:1.5;margin:0 0 4px;">
          Call-ups. Milestones. The weekly roster. Follow us for every move FSU alumni make.
        </p>
        <a href="https://twitter.com/BeyondHowser" target="_blank" rel="noopener"
           style="display:flex;align-items:center;gap:10px;background:#000;color:white;padding:10px 14px;border-radius:8px;text-decoration:none;font-weight:700;font-size:0.85rem;">
          <svg width="18" height="18" viewBox="0 0 24 24" fill="white"><path d="M18.244 2.25h3.308l-7.227 8.26 8.502 11.24H16.17l-4.714-6.231-5.401 6.231H2.746l7.73-8.835L1.254 2.25H8.08l4.254 5.622zm-1.161 17.52h1.833L7.084 4.126H5.117z"/></svg>
          Follow @BeyondHowser
        </a>
        <a href="https://instagram.com/BeyondHowser" target="_blank" rel="noopener"
           style="display:flex;align-items:center;gap:10px;background:linear-gradient(45deg,#f09433,#e6683c,#dc2743,#cc2366,#bc1888);color:white;padding:10px 14px;border-radius:8px;text-decoration:none;font-weight:700;font-size:0.85rem;">
          <svg width="18" height="18" viewBox="0 0 24 24" fill="white"><path d="M12 2.163c3.204 0 3.584.012 4.85.07 3.252.148 4.771 1.691 4.919 4.919.058 1.265.069 1.645.069 4.849 0 3.205-.012 3.584-.069 4.849-.149 3.225-1.664 4.771-4.919 4.919-1.266.058-1.644.07-4.85.07-3.204 0-3.584-.012-4.849-.07-3.26-.149-4.771-1.699-4.919-4.92-.058-1.265-.07-1.644-.07-4.849 0-3.204.013-3.583.07-4.849.149-3.227 1.664-4.771 4.919-4.919 1.266-.057 1.645-.069 4.849-.069zm0-2.163c-3.259 0-3.667.014-4.947.072-4.358.2-6.78 2.618-6.98 6.98-.059 1.281-.073 1.689-.073 4.948 0 3.259.014 3.668.072 4.948.2 4.358 2.618 6.78 6.98 6.98 1.281.058 1.689.072 4.948.072 3.259 0 3.668-.014 4.948-.072 4.354-.2 6.782-2.618 6.979-6.98.059-1.28.073-1.689.073-4.948 0-3.259-.014-3.667-.072-4.947-.196-4.354-2.617-6.78-6.979-6.98-1.281-.059-1.69-.073-4.949-.073zm0 5.838c-3.403 0-6.162 2.759-6.162 6.162s2.759 6.163 6.162 6.163 6.162-2.759 6.162-6.163c0-3.403-2.759-6.162-6.162-6.162zm0 10.162c-2.209 0-4-1.79-4-4 0-2.209 1.791-4 4-4s4 1.791 4 4c0 2.21-1.791 4-4 4zm6.406-11.845c-.796 0-1.441.645-1.441 1.44s.645 1.44 1.441 1.44c.795 0 1.439-.645 1.439-1.44s-.644-1.44-1.439-1.44z"/></svg>
          Follow @BeyondHowser
        </a>
      </div>
    </div>

    <!-- Links -->
    <div class="sidebar-widget">
      <div class="sidebar-widget-hdr">FSU Baseball Links</div>
      <ul class="links-list">
        <li>
          <a href="https://seminoles.com/sports/baseball/" target="_blank" rel="noopener">
            <span class="link-icon" style="background:#782F40">
              <svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
                <circle cx="12" cy="12" r="10" stroke="white" stroke-width="1.5"/>
                <path d="M4.5 9.5c1.5 0 3 .5 4 1.5s2 1.5 3 1.5 2-.5 3-1.5 2.5-1.5 4-1.5" stroke="white" stroke-width="1.5" stroke-linecap="round"/>
                <path d="M4.5 14.5c1.5 0 3-.5 4-1.5s2-1.5 3-1.5 2 .5 3 1.5 2.5 1.5 4 1.5" stroke="white" stroke-width="1.5" stroke-linecap="round"/>
                <line x1="12" y1="2" x2="12" y2="22" stroke="white" stroke-width="1.5" stroke-linecap="round"/>
              </svg>
            </span>
            <span><span class="link-lbl">FSU Baseball</span><span class="link-sub">seminoles.com — Official site</span></span>
          </a>
        </li>
        <li>
          <a href="https://x.com/FSUBaseball" target="_blank" rel="noopener">
            <span class="link-icon" style="background:#000">
              <svg viewBox="0 0 24 24" fill="white" xmlns="http://www.w3.org/2000/svg">
                <path d="M18.244 2.25h3.308l-7.227 8.26 8.502 11.24H16.17l-4.714-6.231-5.401 6.231H2.746l7.73-8.835L1.254 2.25H8.08l4.253 5.622 5.912-5.622zm-1.161 17.52h1.833L7.084 4.126H5.117z"/>
              </svg>
            </span>
            <span><span class="link-lbl">@FSUBaseball</span><span class="link-sub">Official FSU Baseball on X</span></span>
          </a>
        </li>
      </ul>
    </div>

  </div>
</div>

<!-- About Section -->
<div class="section-wrap" id="about" style="margin-bottom: 48px;">
  <div class="section-title">About Beyond Howser</div>
  <div class="about-box">
    <p><strong>Beyond Howser</strong> is a fan-run tracker dedicated to following Florida State University baseball alumni throughout their professional careers — from rookie ball all the way to the Major Leagues.</p>
    <p>Florida State has one of the most storied baseball programs in the country. Playing out of <strong>Dick Howser Stadium</strong> in     Tallahassee, FL, the Seminoles have produced dozens of professional players at every level of the game.</p>
    <p>This tracker pulls live stats directly from the <strong>MLB Stats API</strong> and refreshes automatically each day during the season. Player cards show current {SEASON} season stats, and the roster covers all known FSU alumni active in affiliated or independent professional baseball.</p>
    <p>Have a player to add or a correction? Reach out on <a href="https://twitter.com/BeyondHowser" target="_blank" rel="noopener" style="color:var(--garnet)">&#120143; @BeyondHowser</a>.</p>
  </div>
</div>

<footer>
  <span>&#169; {SEASON} <strong style="color:var(--gold)">Beyond Howser</strong></span>
  &nbsp;&#183;&nbsp; Fan site — not affiliated with FSU or MLB
  &nbsp;&#183;&nbsp; <a href="https://twitter.com/BeyondHowser" target="_blank" rel="noopener">&#120143; @BeyondHowser</a>
  &nbsp;&#183;&nbsp; Data: MLB Stats API
</footer>

<script>
const MODAL_DATA = {modal_json};

function openModal(idx) {{
  const d = MODAL_DATA[idx];
  if (!d) return;
  document.getElementById('mName').textContent = d.name;
  document.getElementById('mSub').textContent  = d.posTeam;
  const hdr = document.getElementById('mHeader');
  hdr.style.background = d.color;
  hdr.style.color      = d.txt;
  const init = document.getElementById('mPhotoInit');
  init.textContent      = d.initials;
  init.style.background = d.color;
  init.style.color      = d.txt;
  const photo = document.getElementById('mPhoto');
  if (d.photo) {{
    photo.style.display = '';
    photo.onerror = function() {{
      const s = this.src;
      if (!this.dataset.tried1) {{ this.dataset.tried1=1; this.src=s.replace('/67/','/milb/'); }}
      else if (!this.dataset.tried2) {{ this.dataset.tried2=1; this.src='{SHADOW_URL}'; this.onerror=null; }}
    }};
    photo.dataset.tried1 = '';
    photo.dataset.tried2 = '';
    photo.src = d.photo;
  }} else {{ photo.style.display = 'none'; }}
  const draftEl = document.getElementById('mDraft');
  draftEl.textContent = d.draft ? 'Draft: ' + d.draft : '';
  const notesEl = document.getElementById('mNotes');
  if (d.notes) {{ notesEl.textContent = d.notes; notesEl.style.display = ''; }}
  else {{ notesEl.style.display = 'none'; }}
  document.getElementById('mStatsSection').innerHTML   = d.statsHtml;
  document.getElementById('mGameLogSection').innerHTML = d.gameLogHtml || '';
  // Reset to Season Stats tab
  switchTab('stats');
  const milbLink = document.getElementById('mMilbLink');
  if (d.milbUrl) {{ milbLink.href = d.milbUrl; milbLink.style.display = ''; }}
  else {{ milbLink.style.display = 'none'; }}
  document.getElementById('modalOverlay').classList.add('open');
  document.body.style.overflow = 'hidden';
}}

function switchTab(tab) {{
  const statsEl = document.getElementById('mStatsSection');
  const logEl   = document.getElementById('mGameLogSection');
  const tabS    = document.getElementById('tabStats');
  const tabL    = document.getElementById('tabLog');
  if (tab === 'stats') {{
    statsEl.style.display = '';
    logEl.style.display   = 'none';
    tabS.classList.add('active');
    tabL.classList.remove('active');
  }} else {{
    statsEl.style.display = 'none';
    logEl.style.display   = '';
    tabS.classList.remove('active');
    tabL.classList.add('active');
  }}
}}

function closeModal() {{
  document.getElementById('modalOverlay').classList.remove('open');
  document.body.style.overflow = '';
}}

function maybeClose(e) {{
  if (e.target === document.getElementById('modalOverlay')) closeModal();
}}

document.addEventListener('keydown', e => {{ if (e.key === 'Escape') closeModal(); }});

function setView(v) {{
  const grid = document.getElementById('grid');
  const list = document.getElementById('listWrap');
  const btnC = document.getElementById('btnCards');
  const btnL = document.getElementById('btnList');
  if (v === 'cards') {{
    grid.style.display = ''; list.style.display = 'none';
    btnC.classList.add('active'); btnL.classList.remove('active');
  }} else {{
    grid.style.display = 'none'; list.style.display = 'block';
    btnC.classList.remove('active'); btnL.classList.add('active');
  }}
}}

let currentLevel  = 'all';
let currentSearch = '';

function filterLevel(level, btn) {{
  currentLevel = level;
  document.querySelectorAll('.filter-btn').forEach(b => {{
    b.classList.remove('active');
    b.style.background = b.style.borderColor = b.style.color = '';
  }});
  btn.classList.add('active');
  const c = btn.dataset.color || '#782F40';
  btn.style.background = btn.style.borderColor = c;
  btn.style.color = ['AA'].includes(level) ? '#333' : 'white';
  applyFilters();
}}

function applySearch(q) {{
  currentSearch = q.toLowerCase();
  applyFilters();
}}

const navSections = ['home','roster','news','about'];
function updateActiveNav() {{
  const scrollY = window.scrollY + 80;
  let active = 'home';
  navSections.forEach(id => {{
    const el = document.getElementById(id);
    if (el && el.offsetTop <= scrollY) active = id;
  }});
  document.querySelectorAll('.nav-link').forEach(a => {{
    a.classList.toggle('active', a.getAttribute('href') === '#' + active);
  }});
}}
window.addEventListener('scroll', updateActiveNav, {{ passive: true }});

function applyFilters() {{
  document.querySelectorAll('.card').forEach(card => {{
    const lm = currentLevel === 'all' || card.dataset.level === currentLevel;
    const nm = !currentSearch || card.dataset.name.includes(currentSearch);
    card.classList.toggle('hidden', !lm || !nm);
  }});
  document.querySelectorAll('#listBody tr').forEach(row => {{
    const lm = currentLevel === 'all' || row.dataset.level === currentLevel;
    const nm = !currentSearch || row.dataset.name.includes(currentSearch);
    row.classList.toggle('hidden', !lm || !nm);
  }});
}}
</script>
</body>
</html>"""
    # Substitute modal JSON (can't put it in f-string directly due to braces)
    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"  ✓ Dashboard written → {HTML_PATH.name}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*56}")
    print(f"  Beyond Howser — Stats Updater  ({SEASON})")
    print(f"{'='*56}")

    cache  = load_cache()
    roster = read_roster()
    print(f"\n  Roster loaded: {len(roster)} players")

    player_data   = []
    updated_count = 0

    for player in roster:
        name = player.get("name", "")
        if not name:
            continue

        print(f"  • {name} … ", end="", flush=True)

        # Skip API lookups for released/IL/independent players
        level = player.get("level", "")
        if level in NO_STATS_LEVELS:
            print(f"skipped ({level})")
            player_data.append({**player, "stats_fmt": {}, "mlb_id": None, "game_log": []})
            continue

        pid = find_player_id(name, cache)
        if not pid:
            print("no MLB ID found")
            player_data.append({**player, "stats_fmt": {}, "mlb_id": None, "game_log": []})
            continue

        player["mlb_id"] = pid
        stats_raw = get_player_stats(pid, SEASON, player.get("level", ""))
        if not stats_raw:
            print("no stats")
            player_data.append({**player, "stats_fmt": {}, "mlb_id": pid})
            continue

        pitcher = is_pitcher(player.get("position", ""))
        fmt     = format_pitching(stats_raw["pitching"]) if pitcher else format_hitting(stats_raw["hitting"])
        player["stats_fmt"] = fmt
        # Fetch recent game log
        player["game_log"] = get_game_log(pid, SEASON, player.get("level", ""))
        player_data.append(player)
        updated_count += 1
        summary = f"{list(fmt.items())[0][0]}={list(fmt.items())[0][1]}" if fmt else "no stats yet"
        print(f"ok ({summary})")

    save_cache(cache)

    print(f"\n  Updating Excel …")
    update_excel(player_data)

    print(f"  Generating news cards …")
    news      = generate_news_cards(player_data)
    news_cache = load_news_cache()
    save_news_cache(news_cache)

    print(f"  Building HTML dashboard …")
    generate_html(player_data, news)

    print(f"\n  Done\! {updated_count}/{len(roster)} players had live stats.")
    print(f"  Excel     → {EXCEL_PATH.name}")
    print(f"  Dashboard → {HTML_PATH.name}")
    print(f"{'='*56}\n")


if __name__ == "__main__":
    main()
